"""
Client Data Exporter

Exports client master data to CSV and Excel formats.
"""
import csv
from pathlib import Path
from typing import List, Optional
from decimal import Decimal

from ..models.return_data import TaxReturn
from .scoring import ComplexityScorer, ScoreBreakdown


class ClientExporter:
    """Export client data to various formats"""
    
    def __init__(self, config_path: Optional[str] = None):
        self.scorer = ComplexityScorer(config_path)
    
    def to_rows(self, returns: List[TaxReturn]) -> List[dict]:
        """Convert tax returns to flat row dictionaries"""
        rows = []
        
        for tr in returns:
            score = self.scorer.score(tr)
            
            row = {
                # Client Info
                'client_id': tr.client_id,
                'taxpayer_name': tr.taxpayer.full_name,
                'taxpayer_ssn': tr.taxpayer.ssn,
                'spouse_name': tr.spouse.full_name if tr.spouse else '',
                'filing_status': tr.filing_status.name,
                
                # Contact
                'address': str(tr.address),
                'phone': tr.taxpayer.phone,
                'email': tr.taxpayer.email,
                
                # Income Summary
                'total_wages': float(tr.income.total_wages),
                'total_interest': float(tr.income.total_interest),
                'total_dividends': float(tr.income.total_dividends),
                'total_retirement': float(tr.income.total_retirement_distributions),
                'total_self_employment': float(tr.income.total_self_employment),
                'total_k1_income': float(tr.income.total_k1_income),
                'total_social_security': float(tr.income.total_social_security),
                'total_other_income': float(tr.income.total_other_income),
                
                # Form Counts
                'w2_count': len(tr.income.w2s),
                'k1_count': len(tr.income.k1_1065) + len(tr.income.k1_1120s),
                'fbar_count': len(tr.income.fbar),
                'brokerage_count': len(tr.raw_forms.get('881', [])),
                
                # Scoring
                'complexity_score': score.total_score,
                'document_count': score.document_count,
                'fee_tier': score.fee_tier,
                'suggested_fee': score.suggested_fee,
                
                # Flags
                'has_business': 1 if tr.balance_sheet else 0,
                'has_fbar': 1 if tr.income.fbar else 0,
                'is_mfj': 1 if tr.filing_status.name == 'MARRIED_FILING_JOINTLY' else 0,
            }
            rows.append(row)
        
        return rows
    
    def to_csv(self, returns: List[TaxReturn], output_path: str) -> None:
        """Export to CSV file"""
        rows = self.to_rows(returns)
        if not rows:
            return
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
    
    def to_excel(self, returns: List[TaxReturn], output_path: str) -> None:
        """Export to Excel file with formatting"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            print("openpyxl not installed. Run: pip install openpyxl")
            return
        
        rows = self.to_rows(returns)
        if not rows:
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Client Master"
        
        # Headers
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[key])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_path)


def export_client_master(returns: List[TaxReturn], output_dir: str, 
                         config_path: Optional[str] = None) -> dict:
    """
    Export client master to both CSV and Excel.
    Returns dict with output file paths.
    """
    exporter = ClientExporter(config_path)
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    csv_path = output_path / "client_master.csv"
    xlsx_path = output_path / "client_master.xlsx"
    
    exporter.to_csv(returns, str(csv_path))
    exporter.to_excel(returns, str(xlsx_path))
    
    return {
        'csv': str(csv_path),
        'excel': str(xlsx_path),
        'count': len(returns)
    }
