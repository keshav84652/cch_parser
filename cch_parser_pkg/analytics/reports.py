"""
Reports Module - Save analytics to files

Exports opportunities, risks, K-1 network, and executive summary to files.
"""
import csv
from pathlib import Path
from typing import List, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .opportunities import ClientOpportunities
from .risks import ClientRisks
from .network import K1NetworkAnalysis
from .scoring import ScoreBreakdown


def save_opportunities_report(opportunities: List[ClientOpportunities], 
                               output_dir: str) -> str:
    """Save opportunities to Excel/CSV"""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    # Flatten opportunities
    rows = []
    for client_opp in opportunities:
        for opp in client_opp.opportunities:
            rows.append({
                'Client': client_opp.client_name,
                'Client ID': client_opp.client_id,
                'AGI': float(client_opp.agi),
                'Opportunity': opp.name,
                'Description': opp.description,
                'Est. Revenue': opp.estimated_revenue,
                'Priority': opp.priority,
                'Trigger': opp.trigger_reason
            })
    
    # Save CSV
    csv_path = output_path / "opportunities_report.csv"
    if rows:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
    
    # Save Excel if openpyxl available
    if HAS_OPENPYXL and rows:
        xlsx_path = output_path / "opportunities_report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Opportunities"
        
        # Headers
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[key])
        
        wb.save(xlsx_path)
        return str(xlsx_path)
    
    return str(csv_path)


def save_risks_report(risks: List[ClientRisks], output_dir: str) -> str:
    """Save risk flags to Excel/CSV"""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    rows = []
    for client_risk in risks:
        for flag in client_risk.flags:
            rows.append({
                'Client': client_risk.client_name,
                'Client ID': client_risk.client_id,
                'Category': flag.category,
                'Severity': flag.severity,
                'Description': flag.description,
                'Recommendation': flag.recommendation,
                'Needs Partner Review': 'Yes' if client_risk.needs_partner_review else 'No'
            })
    
    # Save CSV
    csv_path = output_path / "risk_flags_report.csv"
    if rows:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
    
    # Save Excel
    if HAS_OPENPYXL and rows:
        xlsx_path = output_path / "risk_flags_report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risk Flags"
        
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[key])
        
        wb.save(xlsx_path)
        return str(xlsx_path)
    
    return str(csv_path)


def save_k1_network_report(network: K1NetworkAnalysis, output_dir: str) -> str:
    """Save K-1 network analysis to Excel/CSV"""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    # Sheet 1: Top partnerships
    partnership_rows = []
    for p in network.top_partnerships:
        partnership_rows.append({
            'Partnership': p.name,
            'Client Count': p.client_count,
            'Total Income': p.total_income,
            'Clients': ', '.join(p.clients[:5]) + ('...' if len(p.clients) > 5 else '')
        })
    
    # Sheet 2: Shared partnerships only
    shared_rows = []
    for p in network.shared_partnerships:
        for client in p.clients:
            shared_rows.append({
                'Partnership': p.name,
                'Client': client,
                'Total Clients': p.client_count
            })
    
    # Save CSV (partnerships)
    csv_path = output_path / "k1_network_report.csv"
    if partnership_rows:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=partnership_rows[0].keys())
            writer.writeheader()
            writer.writerows(partnership_rows)
    
    # Save Excel with multiple sheets
    if HAS_OPENPYXL:
        xlsx_path = output_path / "k1_network_report.xlsx"
        wb = openpyxl.Workbook()
        
        # Sheet 1: Top Partnerships
        ws1 = wb.active
        ws1.title = "Top Partnerships"
        if partnership_rows:
            headers = list(partnership_rows[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            for row_idx, row in enumerate(partnership_rows, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=row[key])
        
        # Sheet 2: Shared Partnerships Detail
        ws2 = wb.create_sheet("Shared Partnerships")
        if shared_rows:
            headers = list(shared_rows[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            for row_idx, row in enumerate(shared_rows, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=row[key])
        
        wb.save(xlsx_path)
        return str(xlsx_path)
    
    return str(csv_path)


def save_executive_summary(
    client_count: int,
    scores: List[ScoreBreakdown],
    opportunities: List[ClientOpportunities],
    risks: List[ClientRisks],
    network: K1NetworkAnalysis,
    top_income: list,
    top_complexity: list,
    output_dir: str
) -> str:
    """Save executive summary to markdown file"""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    summary_path = output_path / "executive_summary.md"
    
    # Calculate metrics
    total_fee = sum(s.suggested_fee for s in scores)
    total_opp_revenue = sum(o.total_potential_revenue for o in opportunities)
    opp_count = sum(len(o.opportunities) for o in opportunities)
    risk_count = sum(len(r.flags) for r in risks)
    high_risk = sum(1 for r in risks if r.needs_partner_review)
    
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write(f"# Tax Practice Executive Summary\n\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
        
        f.write("## Key Metrics\n\n")
        f.write(f"| Metric | Value |\n")
        f.write(f"|--------|-------|\n")
        f.write(f"| Total Clients | {client_count} |\n")
        f.write(f"| Base Revenue | ${total_fee:,} |\n")
        f.write(f"| Opportunity Revenue | ${total_opp_revenue:,} |\n")
        f.write(f"| **Combined Potential** | **${total_fee + total_opp_revenue:,}** |\n")
        f.write(f"| Service Opportunities | {opp_count} |\n")
        f.write(f"| Risk Flags | {risk_count} |\n")
        f.write(f"| Clients Needing Review | {high_risk} |\n")
        f.write(f"| Total K-1s | {network.total_k1s} |\n")
        f.write(f"| Unique Partnerships | {network.unique_partnerships} |\n")
        f.write(f"| Shared Partnerships | {len(network.shared_partnerships)} |\n\n")
        
        f.write("## Top 10 Clients by Income\n\n")
        f.write("| Rank | Client | Income |\n")
        f.write("|------|--------|--------|\n")
        for i, (name, income) in enumerate(top_income[:10], 1):
            f.write(f"| {i} | {name} | ${income:,.0f} |\n")
        
        f.write("\n## Top 10 Clients by Complexity\n\n")
        f.write("| Rank | Client | Score | Fee |\n")
        f.write("|------|--------|-------|-----|\n")
        for i, (name, score, fee) in enumerate(top_complexity[:10], 1):
            f.write(f"| {i} | {name} | {score} | ${fee:,} |\n")
        
        f.write("\n## Clients Requiring Partner Review\n\n")
        reviewed = [r for r in risks if r.needs_partner_review]
        for r in reviewed:
            f.write(f"- **{r.client_name}**\n")
            for flag in r.flags:
                if flag.severity == 'High':
                    f.write(f"  - {flag.description}\n")
        
        f.write("\n---\n")
        f.write("*See detailed reports in opportunities_report.xlsx, risk_flags_report.xlsx, k1_network_report.xlsx*\n")
    
    return str(summary_path)
