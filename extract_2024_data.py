
import pandas as pd
import os
import openpyxl

# Paths
source_dir = r'c:\Users\kesha\Downloads\taxdocchecklistincomerecofortheyear2024'
output_dir = r'c:\Users\kesha\OneDrive\cch_parser\output\extracted_2024_checklists'

if not os.path.exists(output_dir):
    os.makedirs(output_dir)

files = [f for f in os.listdir(source_dir) if f.endswith('.xlsx')]

print(f"Found {len(files)} Excel files.")

for file in files:
    print(f"\nProcessing: {file}")
    file_path = os.path.join(source_dir, file)
    try:
        # Load workbook to check sheet names and hidden rows
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # We are interested in '2024' tab primarily, but will list all
        target_sheets = [s for s in wb.sheetnames if '2024' in s]
        if not target_sheets:
            print("  No '2024' tab found. Extracting first sheet as fallback.")
            target_sheets = [wb.sheetnames[0]]
        
        for sheet_name in target_sheets:
            print(f"  Extracting Sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            extracted_rows = []
            
            # Iterate rows
            for row in ws.iter_rows(values_only=True):
                # Filter None/Empty
                clean_row = [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != '']
                if clean_row:
                    extracted_rows.append(clean_row)
            
            if extracted_rows:
                # Save to CSV
                # Create a DataFrame for nice CSV formatting (though structure is ragged)
                # We will just write a ragged CSV manually or use pandas with max columns
                max_cols = max(len(r) for r in extracted_rows)
                # Pad rows
                padded_rows = [r + [''] * (max_cols - len(r)) for r in extracted_rows]
                
                df = pd.DataFrame(padded_rows)
                # Clean filename
                clean_name = os.path.splitext(file)[0].replace(' ', '_')
                clean_sheet = sheet_name.replace(' ', '_')
                out_name = f"{clean_name}_{clean_sheet}.csv"
                out_path = os.path.join(output_dir, out_name)
                
                df.to_csv(out_path, index=False, header=False)
                print(f"    Saved to: {out_name}")
            else:
                print("    Sheet was empty (no data found).")
                
    except Exception as e:
        print(f"  Error processing file: {e}")

print("\nExtraction Complete.")
